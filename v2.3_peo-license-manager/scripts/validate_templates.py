#!/usr/bin/env python3
"""
validate_templates.py  –  PEO License Manager  v2.0
-----------------------------------------------------
Reads every PDF template in templates/pdf/, extracts its AcroForm field names,
compares them against the Fields_Templates sheet baseline in registros.xlsm,
and writes report-validation.json.

The HTML control station reads report-validation.json on folder load to display
red alert dots on states where the template has changed.

Run weekly (Endpoint Central scheduled task):
    python scripts/validate_templates.py

Or on demand:
    python scripts/validate_templates.py --force-baseline
"""

import hashlib
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    from pypdf import PdfReader
except ImportError:
    print("pypdf not installed. Run:  pip install pypdf")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent.parent
WORKBOOK      = BASE_DIR / "data" / "registros.xlsm"
TEMPLATES_DIR = BASE_DIR / "templates" / "pdf"
BASELINE      = BASE_DIR / "field_hashes.json"
REPORT_OUT    = BASE_DIR / "report-validation.json"
LOG_FILE      = BASE_DIR / "logs" / "validate_log.txt"

FIELDS_SHEET  = "Fields_Templates"

# ── Logging ────────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── PDF helpers ────────────────────────────────────────────────────────────
def extract_pdf_fields(pdf_path: Path) -> dict:
    """
    Return a dict of {field_name: field_type} for all AcroForm fields.
    Field types: /Tx (text), /Btn (checkbox/radio), /Ch (dropdown).
    """
    reader = PdfReader(str(pdf_path))
    fields = {}
    if reader.get_fields():
        for name, info in reader.get_fields().items():
            ft = info.get("/FT", "unknown")
            if hasattr(ft, "strip"):
                ft = ft.strip("/")
            fields[name] = str(ft)
    return fields


def field_fingerprint(fields: dict) -> str:
    """SHA-256 of sorted field names — detects add/remove/rename."""
    key = "|".join(sorted(fields.keys()))
    return hashlib.sha256(key.encode()).hexdigest()[:16]


# ── Workbook helpers ───────────────────────────────────────────────────────
def load_fields_template_mapping(workbook_path: Path) -> dict:
    """
    Returns {(estado, proceso): [field_name, ...]} from Fields_Templates sheet.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, keep_vba=False)
    if FIELDS_SHEET not in wb.sheetnames:
        log.warning("Sheet '%s' not found — skipping workbook mapping check.", FIELDS_SHEET)
        return {}

    ws    = wb[FIELDS_SHEET]
    rows  = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(v).strip() for v in rows[0]]
    try:
        col_estado  = headers.index("Estado")
        col_proceso = headers.index("Proceso")
        col_pdf     = headers.index("field_pdf")
    except ValueError as exc:
        log.warning("Fields_Templates missing expected column: %s", exc)
        return {}

    mapping: dict[tuple, list] = {}
    for row in rows[1:]:
        if not any(row):
            continue
        estado  = str(row[col_estado]  or "").strip()
        proceso = str(row[col_proceso] or "").strip()
        field   = str(row[col_pdf]     or "").strip()
        if estado and proceso and field:
            key = (estado, proceso)
            mapping.setdefault(key, []).append(field)

    return mapping


# ── Baseline I/O ───────────────────────────────────────────────────────────
def load_baseline() -> dict:
    if BASELINE.exists():
        with BASELINE.open(encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_baseline(data: dict) -> None:
    with BASELINE.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    log.info("Baseline saved: %s", BASELINE)


# ── Main validation ────────────────────────────────────────────────────────
def validate_all(force_baseline: bool = False) -> list:
    """
    Walk templates/pdf/, check each PDF, compare to baseline.
    Returns a list of result dicts (one per template).
    """
    if not TEMPLATES_DIR.exists():
        log.error("Templates directory not found: %s", TEMPLATES_DIR)
        return []

    baseline     = load_baseline()
    wb_mapping   = load_fields_template_mapping(WORKBOOK) if WORKBOOK.exists() else {}
    results      = []
    new_baseline = {}

    pdf_files = sorted(TEMPLATES_DIR.glob("plantilla_*.pdf"))
    if not pdf_files:
        log.warning("No PDF templates found in %s", TEMPLATES_DIR)
        return []

    for pdf_path in pdf_files:
        # Parse filename: plantilla_{Estado}_{Proceso}.pdf
        stem  = pdf_path.stem          # plantilla_Arkansas_TERM
        parts = stem.split("_", 2)    # ['plantilla', 'Arkansas', 'TERM']  (may be multi-word state)
        if len(parts) < 3:
            log.warning("Unexpected template filename: %s — skipping", pdf_path.name)
            continue

        # Reconstruct: everything between first _ and last _ is the state
        # plantilla_New_York_ADD.pdf  → estado=New York, proceso=ADD
        inner     = "_".join(parts[1:])     # New_York_ADD
        proceso   = inner.rsplit("_", 1)[1] # ADD
        estado    = inner.rsplit("_", 1)[0].replace("_", " ")  # New York

        log.info("Checking: %s  [%s / %s]", pdf_path.name, estado, proceso)

        try:
            fields      = extract_pdf_fields(pdf_path)
            fingerprint = field_fingerprint(fields)
        except Exception as exc:
            log.error("Could not read %s: %s", pdf_path.name, exc)
            results.append({
                "file":        pdf_path.name,
                "estado":      estado,
                "proceso":     proceso,
                "status":      "error",
                "error":       str(exc),
                "checked_at":  datetime.now().isoformat(timespec="seconds"),
            })
            continue

        prev       = baseline.get(pdf_path.name, {})
        prev_fp    = prev.get("fingerprint", "")
        prev_names = set(prev.get("fields", {}).keys())
        curr_names = set(fields.keys())

        added   = sorted(curr_names - prev_names)
        removed = sorted(prev_names - curr_names)
        changed = fingerprint != prev_fp and not force_baseline

        # Cross-check against Fields_Templates workbook mapping
        wb_key     = (estado, proceso)
        wb_fields  = set(wb_mapping.get(wb_key, []))
        not_in_wb  = sorted(curr_names - wb_fields) if wb_fields else []
        not_in_pdf = sorted(wb_fields - curr_names)  if wb_fields else []

        status = "ok"
        if force_baseline:
            status = "baseline_set"
        elif changed:
            status = "changed"
        elif not_in_pdf:
            status = "mapping_mismatch"

        entry = {
            "file":          pdf_path.name,
            "estado":        estado,
            "proceso":       proceso,
            "status":        status,
            "fingerprint":   fingerprint,
            "field_count":   len(curr_names),
            "added_fields":  added,
            "removed_fields": removed,
            "not_in_workbook": not_in_wb,
            "not_in_pdf":    not_in_pdf,
            "checked_at":    datetime.now().isoformat(timespec="seconds"),
        }
        results.append(entry)

        level = logging.WARNING if status not in ("ok", "baseline_set") else logging.INFO
        log.log(level, "  Status: %-20s  Fields: %d  Added: %d  Removed: %d",
                status, len(curr_names), len(added), len(removed))

        new_baseline[pdf_path.name] = {
            "fingerprint": fingerprint,
            "fields":      fields,
            "last_seen":   datetime.now().isoformat(timespec="seconds"),
        }

    # Save updated baseline
    if force_baseline or not baseline:
        save_baseline(new_baseline)
        log.info("Baseline set for %d template(s).", len(new_baseline))
    else:
        # Update baseline entries for unchanged templates
        merged = dict(baseline)
        merged.update(new_baseline)
        save_baseline(merged)

    return results


def write_report(results: list) -> None:
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total":        len(results),
        "ok":           sum(1 for r in results if r.get("status") == "ok"),
        "changed":      sum(1 for r in results if r.get("status") == "changed"),
        "errors":       sum(1 for r in results if r.get("status") == "error"),
        "alerts":       [r for r in results if r.get("status") not in ("ok", "baseline_set")],
        "all":          results,
    }
    with REPORT_OUT.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    log.info("report-validation.json written: %d alert(s)", len(summary["alerts"]))


# ── Entry point ────────────────────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser(description="Validate PEO PDF templates against baseline")
    parser.add_argument(
        "--force-baseline",
        action="store_true",
        help="Re-set the baseline from current templates (use after intentional form update)"
    )
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("PEO Template Validator  |  %s", datetime.now().strftime("%Y-%m-%d %H:%M"))
    log.info("Force baseline: %s", args.force_baseline)
    log.info("=" * 60)

    results = validate_all(force_baseline=args.force_baseline)
    write_report(results)

    alerts = [r for r in results if r.get("status") not in ("ok", "baseline_set")]
    if alerts:
        log.warning("%d template alert(s) found — check report-validation.json", len(alerts))
        for a in alerts:
            log.warning("  !! %s  [%s]", a["file"], a["status"])
    else:
        log.info("All templates match baseline.")

    log.info("=" * 60)


if __name__ == "__main__":
    main()
