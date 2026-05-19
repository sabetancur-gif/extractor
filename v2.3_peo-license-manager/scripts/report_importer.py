#!/usr/bin/env python3
"""
report_importer.py  –  PEO License Manager  v2.0
--------------------------------------------------
Reads the weekly compliance report (Excel) and appends new records
to registros.xlsm > Data sheet.

Also writes session.json so the HTML app knows the current user and
the last import timestamp.

Run:
    python scripts/report_importer.py --report path/to/report.xlsx

Endpoint Central deployment:
    python scripts/report_importer.py --report "%USERPROFILE%\\Downloads\\report.xlsx"
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent.parent
WORKBOOK    = BASE_DIR / "data" / "registros.xlsm"
CONFIG      = BASE_DIR / "report_config.json"
SESSION     = BASE_DIR / "session.json"
KPI_LOG     = BASE_DIR / "logs" / "kpi_events.json"
LOG_FILE    = BASE_DIR / "logs" / "import_log.txt"

DATA_SHEET          = "Data"
FIELDS_SHEET        = "Fields_Templates"
KPI_SHEET           = "KPI"

# ── Logging ────────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── Helpers ────────────────────────────────────────────────────────────────
def norm(v: str) -> str:
    """Normalise a string for fuzzy column matching."""
    return str(v or "").strip().lower().replace(" ", "_").replace("\\", "/")


def get_username() -> str:
    """Return the OS username (works on Windows and macOS/Linux)."""
    return (
        os.environ.get("USERNAME")          # Windows
        or os.environ.get("USER")           # macOS / Linux
        or os.environ.get("LOGNAME")        # fallback
        or "unknown"
    )


def load_config() -> dict:
    if not CONFIG.exists():
        log.error("report_config.json not found at %s", CONFIG)
        log.error("Copy report_config.template.json to report_config.json and fill it in.")
        sys.exit(1)
    with CONFIG.open(encoding="utf-8") as f:
        return json.load(f)


def load_kpi_log() -> list:
    if KPI_LOG.exists():
        with KPI_LOG.open(encoding="utf-8") as f:
            return json.load(f)
    return []


def save_kpi_log(events: list) -> None:
    KPI_LOG.parent.mkdir(parents=True, exist_ok=True)
    with KPI_LOG.open("w", encoding="utf-8") as f:
        json.dump(events, f, indent=2, default=str)


# ── Core logic ─────────────────────────────────────────────────────────────
def read_report(report_path: Path, cfg: dict) -> pd.DataFrame:
    """
    Read the source report and return a DataFrame with column names
    normalised to match the Data sheet headers.
    """
    sheet = cfg.get("report_sheet", 0)          # sheet name or 0-based index
    skip  = cfg.get("report_header_row", 0)     # rows to skip before header
    log.info("Reading report: %s (sheet=%s, skip=%s)", report_path, sheet, skip)

    df = pd.read_excel(report_path, sheet_name=sheet, skiprows=skip, dtype=str)
    df.fillna("", inplace=True)

    # Apply column mapping from config:
    # "column_map": { "Report Column Name": "Data sheet header" }
    col_map: dict = cfg.get("column_map", {})
    df.rename(columns=col_map, inplace=True)

    # Hard-coded state and process overrides (optional)
    if "state_override" in cfg:
        df["Estado"] = cfg["state_override"]
    if "process_override" in cfg:
        df["Proceso"] = cfg["process_override"]

    # Filter to only rows we can use (must have Estado, Proceso, # Case)
    required = ["Estado", "Proceso", "# Case"]
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        log.error("Required columns missing after mapping: %s", missing_cols)
        log.error("Check report_config.json column_map.")
        sys.exit(1)

    df = df[df["# Case"].str.strip() != ""]
    log.info("Report rows after filtering: %d", len(df))
    return df


def get_existing_cases(ws) -> set:
    """Return the set of # Case values already in the Data sheet."""
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    try:
        case_idx = next(
            i for i, h in enumerate(headers)
            if norm(str(h)) == norm("# Case")
        )
    except StopIteration:
        return set()
    cases = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[case_idx]
        if val is not None and str(val).strip():
            cases.add(str(val).strip())
    return cases


def get_data_headers(ws) -> list:
    return [cell.value for cell in next(ws.iter_rows(max_row=1))]


def append_to_workbook(new_rows: pd.DataFrame, workbook_path: Path, username: str) -> tuple[int, int]:
    """
    Append new rows to the Data sheet.
    Returns (appended_count, skipped_count).
    """
    log.info("Opening workbook: %s", workbook_path)
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)

    if DATA_SHEET not in wb.sheetnames:
        log.error("Sheet '%s' not found in workbook.", DATA_SHEET)
        sys.exit(1)

    ws = wb[DATA_SHEET]
    existing = get_existing_cases(ws)
    headers  = get_data_headers(ws)

    appended = 0
    skipped  = 0
    now_str  = datetime.now().strftime("%Y-%m-%d %H:%M")

    for _, report_row in new_rows.iterrows():
        case = str(report_row.get("# Case", "")).strip()
        if not case:
            skipped += 1
            continue
        if case in existing:
            log.debug("Skipping duplicate case: %s", case)
            skipped += 1
            continue

        # Build the row in header order; unknown columns → ""
        row_values = []
        for h in headers:
            if norm(str(h)) == norm("Creado"):
                row_values.append("")           # always blank on import
            elif norm(str(h)) == norm("Imported_By"):
                row_values.append(username)
            elif norm(str(h)) == norm("Import_Date"):
                row_values.append(now_str)
            else:
                row_values.append(report_row.get(h, ""))

        ws.append(row_values)
        existing.add(case)
        appended += 1
        log.info("  + Added case %s (%s / %s)",
                 case,
                 report_row.get("Estado", ""),
                 report_row.get("Proceso", ""))

    wb.save(workbook_path)
    log.info("Workbook saved. Appended: %d  Skipped: %d", appended, skipped)
    return appended, skipped


def ensure_kpi_sheet(workbook_path: Path) -> None:
    """Add a KPI sheet to the workbook if it doesn't exist yet."""
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    if KPI_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(KPI_SHEET)
        ws.append([
            "Timestamp", "Username", "Machine", "Action",
            "State", "Process", "Case", "File", "Duration_sec"
        ])
        wb.save(workbook_path)
        log.info("Created '%s' sheet in workbook.", KPI_SHEET)


def write_session(username: str) -> None:
    data = {
        "username":    username,
        "machine":     os.environ.get("COMPUTERNAME", os.environ.get("HOSTNAME", "unknown")),
        "last_import": datetime.now().isoformat(timespec="seconds"),
        "workbook":    str(WORKBOOK),
    }
    with SESSION.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    log.info("session.json written for user: %s", username)


def log_kpi_event(
    username: str,
    action: str,
    state: str = "",
    process: str = "",
    case: str = "",
    file: str = "",
    duration_sec: float = 0,
) -> None:
    """Append a KPI event to kpi_events.json and to the workbook KPI sheet."""
    events = load_kpi_log()
    event = {
        "timestamp":    datetime.now().isoformat(timespec="seconds"),
        "username":     username,
        "machine":      os.environ.get("COMPUTERNAME", os.environ.get("HOSTNAME", "unknown")),
        "action":       action,
        "state":        state,
        "process":      process,
        "case":         case,
        "file":         file,
        "duration_sec": round(duration_sec, 2),
    }
    events.append(event)
    save_kpi_log(events)

    # Also write to workbook KPI sheet
    if WORKBOOK.exists():
        try:
            wb = openpyxl.load_workbook(WORKBOOK, keep_vba=True)
            if KPI_SHEET in wb.sheetnames:
                ws = wb[KPI_SHEET]
                ws.append([
                    event["timestamp"], event["username"], event["machine"],
                    event["action"],    event["state"],    event["process"],
                    event["case"],      event["file"],     event["duration_sec"],
                ])
                wb.save(WORKBOOK)
        except Exception as exc:
            log.warning("Could not write KPI to workbook: %s", exc)


# ── Entry point ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Import compliance report into registros.xlsm")
    parser.add_argument("--report", required=True, help="Path to the downloaded report Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without saving")
    args = parser.parse_args()

    report_path = Path(args.report)
    if not report_path.exists():
        log.error("Report file not found: %s", report_path)
        sys.exit(1)

    username = get_username()
    log.info("=" * 60)
    log.info("PEO Report Importer  |  User: %s  |  %s", username, datetime.now().strftime("%Y-%m-%d %H:%M"))
    log.info("=" * 60)

    cfg      = load_config()
    new_rows = read_report(report_path, cfg)

    if args.dry_run:
        log.info("DRY RUN — no changes will be saved.")
        log.info("Rows that would be imported:")
        for _, row in new_rows.iterrows():
            log.info("  %s  |  %s  |  %s", row.get("# Case",""), row.get("Estado",""), row.get("Proceso",""))
        return

    # Ensure KPI infrastructure exists
    ensure_kpi_sheet(WORKBOOK)

    start = datetime.now()
    appended, skipped = append_to_workbook(new_rows, WORKBOOK, username)
    elapsed = (datetime.now() - start).total_seconds()

    write_session(username)
    log_kpi_event(
        username=username,
        action="IMPORT",
        file=report_path.name,
        duration_sec=elapsed,
    )

    log.info("Done. Appended: %d  Skipped (duplicates): %d  Time: %.1fs", appended, skipped, elapsed)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
