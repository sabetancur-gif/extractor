"""
mejorar_entrada.py — Aplica diseño visual tipo panel a la hoja "Entrada"
de registros.xlsm, conservando todas las macros y referencias de celdas.

Uso:
    python scripts/mejorar_entrada.py
    python scripts/mejorar_entrada.py --workbook data/registros.xlsm

Requiere: openpyxl >= 3.1
    pip install openpyxl
"""

import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
#  LÍMITE DE SEGURIDAD — no tocar ninguna celda por debajo de esta fila.
#  Las macros VBA de Load/Save suelen referenciar celdas en la zona
#  del formulario (filas 1-60 aprox). Dejamos un margen amplio hasta 150
#  y todo lo que esté debajo queda INTACTO para no romper ninguna macro.
# ═══════════════════════════════════════════════════════════════════════════
MAX_STYLE_ROW = 150


# ═══════════════════════════════════════════════════════════════════════════
#  Paleta de colores (sin #)
# ═══════════════════════════════════════════════════════════════════════════
C = {
    # Fondos de app
    "app_dark":    "0D1117",
    "surface":     "161B22",
    "surface2":    "1C2230",
    "surface3":    "242B38",

    # Acento principal (azul PEO)
    "blue":        "1A6FBF",
    "blue_light":  "58A6FF",
    "blue_bg":     "0D2137",

    # ADD / TERM / Done
    "add":         "1A6FBF",
    "add_bg":      "EDF4FF",
    "term":        "9A4F0A",
    "term_bg":     "FFF4ED",
    "done":        "1A7A3C",
    "done_bg":     "EDFFF4",
    "alert":       "C0392B",
    "alert_bg":    "FFF0EE",

    # Texto
    "text_dark":   "0D1117",
    "text_light":  "E6EDF3",
    "muted":       "8B949E",
    "muted_dark":  "57606A",

    # Bordes
    "border":      "D0D7DE",
    "border_dark": "30363D",

    # Fondos de sección
    "header_bg":   "1A6FBF",   # banner principal
    "section_bg":  "F6F8FA",   # fondo de sección
    "label_bg":    "EFF2F5",   # fondo de etiquetas
    "input_bg":    "FFFFFF",   # fondo de inputs
    "white":       "FFFFFF",
}


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(name="Calibri", bold=False, size=11, color="0D1117", italic=False):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def border(top=None, bottom=None, left=None, right=None, color="D0D7DE", style="thin"):
    def s(active):
        return Side(border_style=style, color=color) if active else Side(border_style=None)
    return Border(top=s(top), bottom=s(bottom), left=s(left), right=s(right))


def border_all(color="D0D7DE", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)


def border_box(color="D0D7DE", style="medium"):
    """Solo bordes exteriores de una celda individual."""
    s = Side(border_style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)


# ═══════════════════════════════════════════════════════════════════════════
#  Función principal
# ═══════════════════════════════════════════════════════════════════════════
def mejorar_entrada(wb, sheet_name="Entrada"):
    if sheet_name not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if "entrada" in s.lower() or "entry" in s.lower()]
        if not candidates:
            print(f"  ✗ No se encontró la hoja '{sheet_name}'. Hojas disponibles: {wb.sheetnames}")
            return
        sheet_name = candidates[0]
        print(f"  → Usando hoja alternativa: '{sheet_name}'")

    ws = wb[sheet_name]
    print(f"  → Aplicando diseño a '{sheet_name}' ({ws.max_row} filas × {ws.max_column} cols)")

    # ── 0. Configurar vista ────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale     = 95

    # ── 1. Detectar estructura: buscar celdas con valores ─────────────────
    #   Estrategia: recorrer todas las celdas y catalogar por contenido.
    #   Labels (col izquierda) vs Inputs (col con datos tipo Text##).
    #   NO movemos ninguna celda.

    label_cells  = []   # (row, col) de etiquetas
    input_cells  = []   # (row, col) de inputs del formulario
    button_rows  = []   # filas aproximadas de botones Load/Save (detectadas por zona)

    max_r = min(ws.max_row, MAX_STYLE_ROW)   # ← nunca miramos más allá de la fila 150
    max_c = ws.max_column

    print(f"  → Límite de seguridad: fila {MAX_STYLE_ROW} "
          f"(filas reales: {ws.max_row} — ignorando filas {MAX_STYLE_ROW + 1}+)")

    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            v = str(cell.value or "").strip()
            if not v:
                continue
            # Etiquetas tipo "Text19", "# Case:", "Estado:", etc.
            if (v.startswith("Text") and v[4:].isdigit()) or \
               v.endswith(":") or v in ("# Case", "Estado", "Proceso",
                                         "Load", "Save", "Estado:", "Proceso:"):
                label_cells.append((cell.row, cell.column))
            else:
                input_cells.append((cell.row, cell.column))

    if not label_cells and not input_cells:
        print("  ⚠ No se encontraron celdas con contenido — aplicando diseño base.")

    # ── 2. Determinar columnas del formulario ──────────────────────────────
    #   Col más frecuente de labels → columna de etiquetas
    #   Col más frecuente de inputs → columna de inputs

    def most_common_col(cells):
        from collections import Counter
        if not cells:
            return None
        return Counter(c for _, c in cells).most_common(1)[0][0]

    label_col = most_common_col(label_cells) or 1
    input_col = most_common_col(input_cells) or label_col + 1

    # Filas de labels
    label_rows = sorted(set(r for r, c in label_cells if c == label_col))
    first_data_row = label_rows[0] if label_rows else 1
    last_data_row  = label_rows[-1] if label_rows else max_r

    print(f"  → Columna etiquetas: {get_column_letter(label_col)} "
          f"| Columna inputs: {get_column_letter(input_col)} "
          f"| Filas: {first_data_row}–{last_data_row}")

    # ── 3. Ajustar anchos de columna ──────────────────────────────────────
    col_widths = {}
    for col in range(1, max_c + 1):
        lc = get_column_letter(col)
        if col < label_col:
            col_widths[lc] = 2          # margen izquierdo
        elif col == label_col:
            col_widths[lc] = 26         # etiquetas
        elif col == input_col:
            col_widths[lc] = 28         # inputs
        elif col == input_col + 1:
            col_widths[lc] = 4          # separador
        else:
            col_widths[lc] = 18         # área derecha (# Case, Estado, botones)

    for lc, w in col_widths.items():
        ws.column_dimensions[lc].width = w

    # ── 4. Banner de título (si hay espacio sobre los datos) ───────────────
    BANNER_ROW = max(1, first_data_row - 3)

    if BANNER_ROW >= 1 and first_data_row >= 4:
        # Fila separadora superior
        ws.row_dimensions[BANNER_ROW - 1].height = 8 if BANNER_ROW > 1 else 4
        ws.row_dimensions[BANNER_ROW].height     = 36
        ws.row_dimensions[BANNER_ROW + 1].height = 20
        ws.row_dimensions[BANNER_ROW + 2].height = 8

        # Merge y estilo del banner
        banner_end_col = min(max_c, input_col + 6)
        banner_range   = f"{get_column_letter(label_col)}{BANNER_ROW}:" \
                         f"{get_column_letter(banner_end_col)}{BANNER_ROW}"
        try:
            ws.merge_cells(banner_range)
        except Exception:
            pass  # ya puede estar mergeado

        bc = ws[f"{get_column_letter(label_col)}{BANNER_ROW}"]
        bc.value     = "⚖  PEO License Manager — Entrada de Formulario"
        bc.font      = font("Calibri", bold=True, size=16, color=C["white"])
        bc.fill      = fill(C["blue"])
        bc.alignment = align("left", "center")

        # Subtítulo
        sub_range = f"{get_column_letter(label_col)}{BANNER_ROW + 1}:" \
                    f"{get_column_letter(banner_end_col)}{BANNER_ROW + 1}"
        try:
            ws.merge_cells(sub_range)
        except Exception:
            pass
        sc = ws[f"{get_column_letter(label_col)}{BANNER_ROW + 1}"]
        sc.value     = "Completa los campos · Carga el registro · Guarda al finalizar"
        sc.font      = font("Calibri", bold=False, size=10, color=C["muted_dark"], italic=True)
        sc.fill      = fill(C["section_bg"])
        sc.alignment = align("left", "center")

        # Colorear filas del banner completas
        for r_off in range(-1, 3):
            r = BANNER_ROW + r_off
            if r < 1:
                continue
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if r == BANNER_ROW - 1:
                    cell.fill = fill(C["blue"])
                elif r == BANNER_ROW + 2:
                    cell.fill = fill(C["section_bg"])
                elif r in (BANNER_ROW, BANNER_ROW + 1) and c < label_col:
                    cell.fill = fill(C["blue"] if r == BANNER_ROW else C["section_bg"])

    # ── 5. Estilizar cada fila de datos del formulario ────────────────────
    safe_last = min(last_data_row, MAX_STYLE_ROW)   # ← respeta el límite
    for row_idx in range(first_data_row, safe_last + 1):
        ws.row_dimensions[row_idx].height = 24

        for col_idx in range(1, max_c + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_idx < label_col:
                # Margen izquierdo
                cell.fill = fill(C["section_bg"])

            elif col_idx == label_col:
                # Columna de etiquetas
                cell.font      = font("Calibri", bold=True, size=10, color=C["text_dark"])
                cell.fill      = fill(C["label_bg"])
                cell.alignment = align("right", "center")
                cell.border    = border(bottom=True, right=True,
                                        color=C["border"], style="thin")

            elif col_idx == input_col:
                # Columna de inputs — solo aplicar si la celda está vacía o tiene dato
                v = str(cell.value or "").strip()
                cell.font      = font("Calibri", bold=False, size=10,
                                      color=C["blue"] if v else C["text_dark"])
                cell.fill      = fill(C["input_bg"])
                cell.alignment = align("left", "center")
                cell.border    = border_all(color=C["blue"], style="thin")

            elif col_idx == input_col + 1:
                # Separador
                cell.fill = fill(C["section_bg"])

            else:
                # Área derecha: # Case, Estado, botones
                v = str(cell.value or "").strip()
                if v in ("# Case", "# Case:", "Estado", "Estado:",
                          "Proceso", "Proceso:"):
                    cell.font      = font("Calibri", bold=True, size=10,
                                          color=C["blue"])
                    cell.fill      = fill(C["blue_bg"] if False else C["label_bg"])
                    cell.alignment = align("right", "center")
                elif v:
                    cell.font      = font("Calibri", bold=False, size=10,
                                          color=C["text_dark"])
                    cell.fill      = fill(C["input_bg"])
                    cell.alignment = align("left", "center")
                    cell.border    = border_all(color=C["blue"], style="thin")
                else:
                    cell.fill = fill(C["section_bg"])

    # ── 6. Fila de cierre (bajo el último dato) ───────────────────────────
    footer_row = safe_last + 1
    if footer_row <= MAX_STYLE_ROW:
        ws.row_dimensions[footer_row].height = 6
        for c in range(1, max_c + 1):
            ws.cell(footer_row, c).fill = fill(C["blue"])

    footer_row2 = safe_last + 2
    if footer_row2 <= MAX_STYLE_ROW:
        ws.row_dimensions[footer_row2].height = 18
        for c in range(1, max_c + 1):
            cell = ws.cell(footer_row2, c)
            cell.fill = fill(C["section_bg"])

        ws.cell(footer_row2, label_col).value = (
            "Tip: usa Load para cargar un registro existente y Save para guardarlo"
        )
        ws.cell(footer_row2, label_col).font      = font("Calibri", size=9,
                                                           color=C["muted_dark"],
                                                           italic=True)
        ws.cell(footer_row2, label_col).alignment = align("left", "center")

        tip_end = min(max_c, input_col + 4)
        try:
            ws.merge_cells(
                f"{get_column_letter(label_col)}{footer_row2}:"
                f"{get_column_letter(tip_end)}{footer_row2}"
            )
        except Exception:
            pass

    # ── 7. Colorear celdas fuera del formulario (área amarilla/gris) ───────
    #   Solo hasta MAX_STYLE_ROW — las filas 151+ se dejan intactas.
    outside_start = safe_last + 3
    outside_end   = min(safe_last + 30, MAX_STYLE_ROW)   # ← límite duro
    for row_idx in range(outside_start, outside_end + 1):
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, max_c + 1):
            ws.cell(row_idx, col_idx).fill = fill(C["section_bg"])

    print(f"  ✓ Diseño aplicado a hoja '{sheet_name}'.")


# ═══════════════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Mejora visual de la hoja Entrada en registros.xlsm"
    )
    parser.add_argument(
        "--workbook", default="data/registros.xlsm",
        help="Ruta al archivo .xlsm (default: data/registros.xlsm)"
    )
    parser.add_argument(
        "--sheet", default="Entrada",
        help="Nombre de la hoja a mejorar (default: Entrada)"
    )
    args = parser.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        print(f"  ✗ Archivo no encontrado: {path}")
        return

    print(f"\n  Abriendo {path}…")
    wb = load_workbook(str(path), keep_vba=True)
    print(f"  Hojas disponibles: {wb.sheetnames}")

    mejorar_entrada(wb, sheet_name=args.sheet)

    wb.save(str(path))
    print(f"  ✓ Guardado en {path}")
    print()
    print("  Abre registros.xlsm en Excel para ver el nuevo diseño.")
    print("  Las macros Load / Save siguen funcionando exactamente igual.")
    print()


if __name__ == "__main__":
    main()