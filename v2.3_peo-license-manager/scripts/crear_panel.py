"""
crear_panel.py — Agrega una hoja "Panel" visual al inicio de registros.xlsm.

Uso:
    python scripts/crear_panel.py
    python scripts/crear_panel.py --workbook ruta/custom.xlsm

Requiere: openpyxl >= 3.1
    pip install openpyxl
"""

import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Paleta de colores (hex sin #) ──────────────────────────────────────────
C = {
    "brand_dark":  "0D1117",   # fondo principal (dark)
    "brand_blue":  "1A6FBF",   # add / primary accent
    "brand_gold":  "9A4F0A",   # term / warning
    "brand_green": "1A7A3C",   # done / ok
    "brand_red":   "C0392B",   # alert

    "header_bg":   "161B22",   # superficie del header
    "surface":     "1C2230",   # tarjeta
    "surface2":    "242B38",   # tarjeta alternativa
    "border":      "30363D",   # borde sutil

    "white":       "E6EDF3",   # texto principal (dark)
    "muted":       "8B949E",   # texto secundario

    "add_bg":      "0D2137",   # fondo badge add
    "term_bg":     "2A1F07",   # fondo badge term
    "done_bg":     "0A2211",   # fondo badge done
    "alert_bg":    "2C0B0B",   # fondo badge alert
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="E6EDF3", italic=False, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(color="30363D", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color="30363D", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(bottom=s)


def build_panel(wb, data_sheet_name="Data"):
    """Crea o reemplaza la hoja 'Panel' con el dashboard visual."""

    # Eliminar hoja existente si ya existe
    if "Panel" in wb.sheetnames:
        del wb["Panel"]

    ws = wb.create_sheet("Panel", 0)   # primera posición

    # ── Dimensiones de columnas ────────────────────────────────────────────
    col_widths = {
        "A": 3,    # margen izquierdo
        "B": 22,   # etiqueta
        "C": 18,   # valor 1
        "D": 18,   # valor 2
        "E": 18,   # valor 3
        "F": 18,   # valor 4
        "G": 3,    # margen derecho
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Fondo general ──────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    def bg(row, col, color=C["brand_dark"]):
        ws.cell(row=row, column=col).fill = fill(color)

    # ─────────────────────────────────────────────────────────────────────
    # FILA 1–2: Banner de título
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 48

    ws.merge_cells("B2:F2")
    title_cell = ws["B2"]
    title_cell.value        = "⚖  PEO License Manager"
    title_cell.font         = font(bold=True, size=22, color=C["white"], name="Calibri")
    title_cell.fill         = fill(C["brand_blue"])
    title_cell.alignment    = align("left", "center")

    # Margen izquierdo del banner
    for r in range(1, 4):
        ws.cell(r, 1).fill = fill(C["brand_dark"])
        ws.cell(r, 7).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # FILA 3: Subtítulo
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:F3")
    sub = ws["B3"]
    sub.value     = "Panel de Control Semanal  ·  Compliance & License Filing"
    sub.font      = font(size=10, color=C["muted"])
    sub.fill      = fill(C["header_bg"])
    sub.alignment = align("left", "center")

    # ─────────────────────────────────────────────────────────────────────
    # FILA 4: Espacio
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 10
    for c in range(1, 8):
        ws.cell(4, c).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # FILAS 5–11: Tarjetas KPI (fórmulas que leen la hoja Data)
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 14   # separador de sección

    def section_header(row, label, col_start="B", col_end="F"):
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
        c = ws[f"{col_start}{row}"]
        c.value     = f"  {label}"
        c.font      = font(bold=True, size=9, color=C["brand_blue"])
        c.fill      = fill(C["brand_dark"])
        c.alignment = align("left", "center")
        ws.cell(row, 1).fill = fill(C["brand_dark"])
        ws.cell(row, 7).fill = fill(C["brand_dark"])

    section_header(6, "▌ KPIs del Ciclo Actual")

    # Fila 7: encabezados de tarjetas KPI
    ws.row_dimensions[7].height = 18
    kpi_headers = ["Pendientes ADD", "Pendientes TERM", "Generados", "Alertas"]
    kpi_cols    = ["C", "D", "E", "F"]
    for col, hdr in zip(kpi_cols, kpi_headers):
        c = ws[f"{col}7"]
        c.value     = hdr
        c.font      = font(size=9, color=C["muted"], bold=True)
        c.fill      = fill(C["surface"])
        c.alignment = align("center", "center")

    ws["B7"].fill = fill(C["surface"])
    ws.cell(7, 1).fill = fill(C["brand_dark"])
    ws.cell(7, 7).fill = fill(C["brand_dark"])

    # Fila 8: valores KPI (fórmulas COUNTIFS sobre la hoja Data)
    ws.row_dimensions[8].height = 38
    n = data_sheet_name
    kpi_formulas = [
        f'=COUNTIFS(\'{n}\'!E:E,"ADD",\'{n}\'!G:G,"<>SI",\'{n}\'!G:G,"<>si")',   # ADD pendientes (col E=Proceso, col G=Creado)
        f'=COUNTIFS(\'{n}\'!E:E,"TERM",\'{n}\'!G:G,"<>SI",\'{n}\'!G:G,"<>si")',  # TERM pendientes
        f'=COUNTIFS(\'{n}\'!G:G,"SI")+COUNTIFS(\'{n}\'!G:G,"si")',                # generados
        '="-"',                                                                    # alertas (sin formula fácil)
    ]
    kpi_colors  = [C["add_bg"], C["term_bg"], C["done_bg"], C["alert_bg"]]
    kpi_fcolors = [C["brand_blue"], C["brand_gold"], C["brand_green"], C["brand_red"]]

    ws["B8"].fill = fill(C["surface"])
    ws.cell(8, 1).fill = fill(C["brand_dark"])
    ws.cell(8, 7).fill = fill(C["brand_dark"])

    for col, formula, bg_col, fc in zip(kpi_cols, kpi_formulas, kpi_colors, kpi_fcolors):
        c = ws[f"{col}8"]
        c.value     = formula
        c.font      = font(bold=True, size=18, color=fc)
        c.fill      = fill(bg_col)
        c.alignment = align("center", "center")

    # Fila 9: espacio
    ws.row_dimensions[9].height = 6
    for c in range(1, 8): ws.cell(9, c).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # FILAS 10–16: Tabla de estados con pendientes (últimos datos de Data)
    # ─────────────────────────────────────────────────────────────────────
    section_header(10, "▌ Estados con Archivos Pendientes")

    # Encabezados de tabla
    ws.row_dimensions[11].height = 20
    table_headers = ["Estado", "ADD", "TERM", "Total", "Prioridad"]
    table_cols    = ["B", "C", "D", "E", "F"]
    for col, hdr in zip(table_cols, table_headers):
        c = ws[f"{col}11"]
        c.value     = hdr
        c.font      = font(bold=True, size=10, color=C["white"])
        c.fill      = fill(C["header_bg"])
        c.alignment = align("center", "center")
        c.border    = border_bottom(C["brand_blue"], "medium")
    ws.cell(11, 1).fill = fill(C["brand_dark"])
    ws.cell(11, 7).fill = fill(C["brand_dark"])

    # Filas de datos (placeholder — el usuario llena o se generan por macro/Power Query)
    sample_states = [
        ("New York",     "", "", "", ""),
        ("Arkansas",     "", "", "", ""),
        ("California",   "", "", "", ""),
        ("Texas",        "", "", "", ""),
        ("Florida",      "", "", "", ""),
    ]
    alt_colors = [C["surface"], C["surface2"]]
    for i, (state, *rest) in enumerate(sample_states):
        row = 12 + i
        ws.row_dimensions[row].height = 20
        alt = alt_colors[i % 2]
        for j, (col, val) in enumerate(zip(table_cols, [state] + rest)):
            c = ws[f"{col}{row}"]
            c.value     = val if val else ""
            c.font      = font(size=10, color=C["white"] if j == 0 else C["muted"])
            c.fill      = fill(alt)
            c.alignment = align("left" if j == 0 else "center", "center")
        ws.cell(row, 1).fill = fill(C["brand_dark"])
        ws.cell(row, 7).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # FILAS 18–20: Separador + sección de instrucciones
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[17].height = 10
    for c in range(1, 8): ws.cell(17, c).fill = fill(C["brand_dark"])

    section_header(18, "▌ Instrucciones de Uso")

    instructions = [
        ("1. Abre la app web",      "Abre index.html en Chrome/Edge (doble clic)."),
        ("2. Selecciona operador",   "Haz clic en tu nombre: Mateo, Isabella o Paulina."),
        ("3. Carga la carpeta",      "Clic en 'Select project folder' → elige ESTA carpeta."),
        ("4. Revisa el mapa",        "Los estados con archivos pendientes se iluminan."),
        ("5. Preview y genera",      "Haz clic en el estado → Preview → Generate."),
        ("6. Guarda el workbook",    "Presiona 'Save workbook' para marcar los registros."),
    ]
    for i, (paso, desc) in enumerate(instructions):
        row = 19 + i
        ws.row_dimensions[row].height = 20
        ws["B" + str(row)].value     = paso
        ws["B" + str(row)].font      = font(bold=True, size=10, color=C["brand_blue"])
        ws["B" + str(row)].fill      = fill(C["surface"])
        ws["B" + str(row)].alignment = align("left", "center")

        ws.merge_cells(f"C{row}:F{row}")
        ws["C" + str(row)].value     = desc
        ws["C" + str(row)].font      = font(size=10, color=C["muted"])
        ws["C" + str(row)].fill      = fill(C["surface"])
        ws["C" + str(row)].alignment = align("left", "center")

        ws.cell(row, 1).fill = fill(C["brand_dark"])
        ws.cell(row, 7).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # FILAS 25–29: Footer con info de versión y leyenda de colores
    # ─────────────────────────────────────────────────────────────────────
    ws.row_dimensions[25].height = 14
    for c in range(1, 8): ws.cell(25, c).fill = fill(C["brand_dark"])

    section_header(26, "▌ Leyenda de Colores")
    ws.row_dimensions[27].height = 20
    legend = [
        ("C27", C["add_bg"],   C["brand_blue"],  "ADD pendiente"),
        ("D27", C["term_bg"],  C["brand_gold"],  "TERM pendiente"),
        ("E27", C["done_bg"],  C["brand_green"], "Generado (SI)"),
        ("F27", C["alert_bg"], C["brand_red"],   "Alerta"),
    ]
    ws["B27"].fill = fill(C["brand_dark"])
    ws.cell(27, 1).fill = fill(C["brand_dark"])
    ws.cell(27, 7).fill = fill(C["brand_dark"])
    for addr, bg_c, txt_c, label in legend:
        c = ws[addr]
        c.value     = label
        c.font      = font(bold=True, size=9, color=txt_c)
        c.fill      = fill(bg_c)
        c.alignment = align("center", "center")

    ws.row_dimensions[28].height = 8
    for c in range(1, 8): ws.cell(28, c).fill = fill(C["brand_dark"])

    ws.row_dimensions[29].height = 16
    ws.merge_cells("B29:F29")
    footer = ws["B29"]
    footer.value     = "PEO License Manager · v2.2 · No modificar esta hoja manualmente"
    footer.font      = font(size=9, color=C["muted"], italic=True)
    footer.fill      = fill(C["header_bg"])
    footer.alignment = align("center", "center")
    ws.cell(29, 1).fill = fill(C["brand_dark"])
    ws.cell(29, 7).fill = fill(C["brand_dark"])

    # ─────────────────────────────────────────────────────────────────────
    # Rellenar celdas vacías (columna A y G como márgenes)
    # ─────────────────────────────────────────────────────────────────────
    for row in range(1, 31):
        for col in [1, 7]:
            ws.cell(row, col).fill = fill(C["brand_dark"])

    # Congelar filas del encabezado
    ws.freeze_panes = "B7"

    # Proteger la hoja (lectura) — opcional, descomenta si quieres bloquearla
    # ws.protection.sheet = True
    # ws.protection.password = "peo2024"

    print(f"  ✓ Hoja 'Panel' creada con {ws.max_row} filas.")
    return ws


def main():
    parser = argparse.ArgumentParser(description="Agrega hoja Panel a registros.xlsm")
    parser.add_argument(
        "--workbook", default="data/registros.xlsm",
        help="Ruta al archivo .xlsm (default: data/registros.xlsm)"
    )
    args = parser.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        print(f"  ✗ Archivo no encontrado: {path}")
        return

    print(f"  Abriendo {path}…")
    wb = load_workbook(str(path), keep_vba=True)

    print(f"  Hojas existentes: {wb.sheetnames}")
    data_sheet = next((s for s in wb.sheetnames if "data" in s.lower()), wb.sheetnames[-1])
    print(f"  Hoja de datos detectada: '{data_sheet}'")

    build_panel(wb, data_sheet_name=data_sheet)

    wb.save(str(path))
    print(f"  ✓ Guardado en {path}")
    print()
    print("  Abre el archivo en Excel para ver el Panel.")
    print("  Si ves los colores en blanco, ve a: Vista → Colores de tema → Office.")


if __name__ == "__main__":
    main()
