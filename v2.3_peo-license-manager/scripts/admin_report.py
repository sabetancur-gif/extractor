#!/usr/bin/env python3
"""
admin_report.py  –  PEO License Manager  v2.0
----------------------------------------------
Reads kpi_events.json (and optionally the workbook KPI sheet) and produces
an Excel admin report with:
  - Summary sheet: totals by user, state, process, week
  - Detail sheet: every logged event
  - Timing sheet: avg / min / max generation time by state+process
  - Alerts sheet: template change events

Run:
    python scripts/admin_report.py
    python scripts/admin_report.py --out reports/kpi_2026_w17.xlsx
    python scripts/admin_report.py --from 2026-04-01 --to 2026-04-30
"""

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).resolve().parent.parent
KPI_LOG   = BASE_DIR / "logs" / "kpi_events.json"
WORKBOOK  = BASE_DIR / "data" / "registros.xlsm"
REPORT_DEFAULT = BASE_DIR / "logs" / f"admin_report_{date.today().isoformat()}.xlsx"

# ── Style helpers ──────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="EBF0F7")
BOLD         = Font(bold=True)
CENTER       = Alignment(horizontal="center")
THIN         = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_header_row(ws, row_num: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN

def style_data_row(ws, row_num: int, ncols: int, alt: bool = False) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.border = THIN

def autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ── Data loading ───────────────────────────────────────────────────────────
def load_events(from_date: date | None, to_date: date | None) -> list[dict]:
    """Load KPI events from kpi_events.json, optionally date-filtered."""
    if not KPI_LOG.exists():
        print(f"kpi_events.json not found at {KPI_LOG}")
        print("No events to report — run the importer and generate some PDFs first.")
        sys.exit(0)

    with KPI_LOG.open(encoding="utf-8") as f:
        events = json.load(f)

    if from_date:
        events = [e for e in events if e.get("timestamp", "")[:10] >= from_date.isoformat()]
    if to_date:
        events = [e for e in events if e.get("timestamp", "")[:10] <= to_date.isoformat()]

    return events


# ── Sheet builders ─────────────────────────────────────────────────────────
def build_summary(wb: openpyxl.Workbook, events: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A2"

    # ── By user ──
    ws.append(["By user"])
    ws["A1"].font = BOLD
    headers = ["Username", "Machine", "Total actions", "PDFs generated", "Imports run", "Avg gen time (s)"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    user_map: dict[str, dict] = {}
    for ev in events:
        u = ev.get("username", "unknown")
        if u not in user_map:
            user_map[u] = {"machine": ev.get("machine",""), "total":0, "pdfs":0, "imports":0, "times":[]}
        user_map[u]["total"] += 1
        if ev.get("action") == "GENERATE":
            user_map[u]["pdfs"] += 1
            t = ev.get("duration_sec")
            if t:
                user_map[u]["times"].append(float(t))
        if ev.get("action") == "IMPORT":
            user_map[u]["imports"] += 1

    for i, (u, d) in enumerate(sorted(user_map.items()), start=3):
        avg = round(sum(d["times"]) / len(d["times"]), 1) if d["times"] else ""
        ws.append([u, d["machine"], d["total"], d["pdfs"], d["imports"], avg])
        style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    # ── Spacer + By state ──
    base = len(user_map) + 4
    ws.cell(row=base, column=1, value="By state / process")
    ws.cell(row=base, column=1).font = BOLD
    h2 = ["State", "Process", "PDFs generated", "Avg gen time (s)", "Last activity"]
    ws.append(h2)
    style_header_row(ws, base + 1, len(h2))

    sp_map: dict[tuple, dict] = {}
    for ev in events:
        if ev.get("action") != "GENERATE":
            continue
        key = (ev.get("state",""), ev.get("process",""))
        if key not in sp_map:
            sp_map[key] = {"count": 0, "times": [], "last": ""}
        sp_map[key]["count"] += 1
        t = ev.get("duration_sec")
        if t:
            sp_map[key]["times"].append(float(t))
        ts = ev.get("timestamp","")
        if ts > sp_map[key]["last"]:
            sp_map[key]["last"] = ts

    for i, ((state, proc), d) in enumerate(sorted(sp_map.items()), start=base + 2):
        avg = round(sum(d["times"]) / len(d["times"]), 1) if d["times"] else ""
        ws.append([state, proc, d["count"], avg, d["last"][:10]])
        style_data_row(ws, i, len(h2), alt=(i % 2 == 0))

    autofit(ws)


def build_detail(wb: openpyxl.Workbook, events: list[dict]) -> None:
    ws = wb.create_sheet("All Events")
    ws.freeze_panes = "A2"

    headers = ["Timestamp", "Username", "Machine", "Action", "State", "Process", "Case", "File", "Duration (s)"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for i, ev in enumerate(events, start=2):
        ws.append([
            ev.get("timestamp",""),
            ev.get("username",""),
            ev.get("machine",""),
            ev.get("action",""),
            ev.get("state",""),
            ev.get("process",""),
            ev.get("case",""),
            ev.get("file",""),
            ev.get("duration_sec",""),
        ])
        style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    autofit(ws)


def build_timing(wb: openpyxl.Workbook, events: list[dict]) -> None:
    ws = wb.create_sheet("Generation Timing")
    ws.freeze_panes = "A2"

    gen_events = [e for e in events if e.get("action") == "GENERATE"]
    headers = ["State", "Process", "Count", "Min (s)", "Avg (s)", "Max (s)", "Total (s)"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    sp: dict[tuple, list] = {}
    for ev in gen_events:
        key = (ev.get("state",""), ev.get("process",""))
        t   = ev.get("duration_sec")
        if t is not None:
            sp.setdefault(key, []).append(float(t))

    for i, ((state, proc), times) in enumerate(sorted(sp.items()), start=2):
        ws.append([
            state, proc, len(times),
            round(min(times), 2),
            round(sum(times)/len(times), 2),
            round(max(times), 2),
            round(sum(times), 2),
        ])
        style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    autofit(ws)


def build_alerts(wb: openpyxl.Workbook, events: list[dict]) -> None:
    ws = wb.create_sheet("Template Alerts")
    ws.freeze_panes = "A2"

    alert_events = [e for e in events if e.get("action", "").startswith("TEMPLATE_")]
    headers = ["Timestamp", "Username", "State", "Process", "Action", "File"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    if not alert_events:
        ws.append(["No template alert events recorded."])
    else:
        for i, ev in enumerate(alert_events, start=2):
            ws.append([
                ev.get("timestamp",""), ev.get("username",""),
                ev.get("state",""),     ev.get("process",""),
                ev.get("action",""),    ev.get("file",""),
            ])
            style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    autofit(ws)


# ── Entry point ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate PEO admin KPI report")
    parser.add_argument("--out",  default=str(REPORT_DEFAULT), help="Output .xlsx path")
    parser.add_argument("--from", dest="from_date", default=None, help="Start date YYYY-MM-DD")
    parser.add_argument("--to",   dest="to_date",   default=None, help="End date YYYY-MM-DD")
    args = parser.parse_args()

    from_date = date.fromisoformat(args.from_date) if args.from_date else None
    to_date   = date.fromisoformat(args.to_date)   if args.to_date   else None

    events = load_events(from_date, to_date)
    print(f"Loaded {len(events)} event(s)")

    if not events:
        print("No events in range.")
        sys.exit(0)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default Sheet

    build_summary(wb, events)
    build_detail(wb, events)
    build_timing(wb, events)
    build_alerts(wb, events)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Admin report saved: {out_path}")


if __name__ == "__main__":
    main()
